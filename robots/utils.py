import datetime

from django.db.models import Count
from openpyxl import Workbook

from robots.models import Robot


def generate_robot_production_summary(
    from_time: datetime.datetime,
    to_time: datetime.datetime,
) -> Workbook:
    workbook = Workbook()

    data = (
        Robot.objects.filter(created__gte=from_time, created__lte=to_time)
        .values("model", "version")
        .annotate(count=Count("id"))
        .order_by("model", "version")
    )

    if data:
        current_model = None
        for entry in data:
            model = entry["model"]
            version = entry["version"]
            count = entry["count"]

            if model != current_model:
                worksheet = workbook.create_sheet(title=model)
                worksheet.append(["Модель", "Версия", "Количество за неделю"])
                current_model = model

            worksheet.append([model, version, count])
    else:
        worksheet = workbook.create_sheet(title="Have no robots")
        worksheet.append(["Модель", "Версия", "Количество за неделю"])
        worksheet.append(["-", "-", "-"])

    # Delete default empty sheet
    if "Sheet" in workbook.sheetnames:
        workbook.remove(workbook["Sheet"])

    return workbook
